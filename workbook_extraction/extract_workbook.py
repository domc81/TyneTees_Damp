#!/usr/bin/env python3
"""
TTDP Workbook Forensic Extractor
=================================
Extracts EVERYTHING from a Tyne Tees Damp Proofing survey workbook (.xlsm)
into a structured markdown file for use in web application design.

Usage:
    python extract_workbook.py <path_to_xlsm> [output_dir]

Requirements:
    pip install openpyxl oletools

Output:
    <SURVEY_TYPE>_WORKBOOK_ANALYSIS.md in the output directory
"""

import sys
import os
import re
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

try:
    from oletools.olevba import VBA_Parser
    HAS_OLETOOLS = True
except ImportError:
    HAS_OLETOOLS = False
    print("WARNING: oletools not installed. VBA extraction will be skipped.")
    print("  Install with: pip install oletools")


# ============================================================================
# HELPERS
# ============================================================================

def cell_val(ws, row, col, max_len=None):
    """Get cell value as string, optionally truncated."""
    v = ws.cell(row, col).value
    if v is None:
        return ""
    s = str(v).replace("\n", " ↵ ").replace("\r", "")
    if max_len and len(s) > max_len:
        return s[:max_len] + "…"
    return s


def cell_val_full(ws, row, col):
    """Get full cell value with no truncation."""
    return cell_val(ws, row, col, max_len=None)


def safe_md(text):
    """Escape pipe characters for markdown tables."""
    return str(text).replace("|", "\\|").replace("\n", " ")


def detect_survey_type(wb):
    """Detect survey type from workbook content."""
    costing = wb['Costing']
    cell_b16 = str(costing.cell(16, 2).value or "").upper()
    if "DAMP" in cell_b16:
        return "DAMP"
    elif "CONDENSATION" in cell_b16:
        return "CONDENSATION"
    elif "TIMBER" in cell_b16:
        return "TIMBER"
    elif "WOODWORM" in cell_b16:
        return "WOODWORM"
    # Fallback: check sheet names
    sheets = [s.upper() for s in wb.sheetnames]
    if "CELL REFERENCES" in sheets or "FILTER-CHECK" in sheets:
        return "TIMBER"
    if "PARTY WALL LETTER" in sheets:
        return "TIMBER"
    for name in ["DAMP", "CONDENSATION", "TIMBER", "WOODWORM"]:
        if name in cell_b16:
            return name
    return "UNKNOWN"


def find_max_data_col(ws, sample_rows=50):
    """Find the rightmost column that contains data by sampling rows."""
    max_col = 1
    for r in range(1, min(sample_rows + 1, ws.max_row + 1)):
        for c in range(ws.max_column, 0, -1):
            if ws.cell(r, c).value is not None:
                max_col = max(max_col, c)
                break
    return max_col


# ============================================================================
# SECTION 1: WORKBOOK OVERVIEW
# ============================================================================

def extract_overview(wb, filepath):
    lines = []
    lines.append("## SECTION 1: WORKBOOK OVERVIEW\n")
    lines.append(f"**File:** `{os.path.basename(filepath)}`\n")
    lines.append(f"**Extracted:** {datetime.now().isoformat()}\n")
    lines.append(f"**Survey Type:** {detect_survey_type(wb)}\n")
    lines.append(f"**Sheet Count:** {len(wb.sheetnames)}\n")
    lines.append("")
    lines.append("| # | Sheet Name | Rows | Cols | State |")
    lines.append("|---|-----------|------|------|-------|")
    for i, name in enumerate(wb.sheetnames, 1):
        ws = wb[name]
        state = ws.sheet_state if hasattr(ws, 'sheet_state') else "visible"
        lines.append(f"| {i} | {name} | {ws.max_row} | {ws.max_column} | {state} |")
    lines.append("")
    return "\n".join(lines)


# ============================================================================
# SECTION 2: REPORT SHEET — Row-by-Row Map
# ============================================================================

def extract_report_sheet(wb):
    lines = []
    lines.append("## SECTION 2: REPORT SHEET — Complete Row-by-Row Map\n")

    if 'Report' not in wb.sheetnames:
        lines.append("*No Report sheet found.*\n")
        return "\n".join(lines)

    ws = wb['Report']
    max_col = min(ws.max_column, 50)  # Cap at 50 cols to be safe
    lines.append(f"**Total Rows:** {ws.max_row}  |  **Total Cols:** {ws.max_column}\n")

    # First, find which columns have data by scanning
    cols_with_data = set()
    for r in range(1, min(20, ws.max_row + 1)):
        for c in range(1, max_col + 1):
            if ws.cell(r, c).value is not None:
                cols_with_data.add(c)

    lines.append("### Key Column Legend")
    lines.append("- **Col A (1):** Validation X marker — `=IF(AG_row=1,\"X\",\"\")`")
    lines.append("- **Col B (2):** Visibility flag — 1=prints, 0=hidden, formula=conditional")
    lines.append("- **Col D (4):** Main content/question text")
    lines.append("- **Col E-Q:** Data entry cells (varies by row)")
    lines.append("- **Col Y (25):** Room show/hide toggle (Timber only)")
    lines.append("- **Col AB/AC (28+):** Surveyor guidance comments")
    lines.append("- **Col AG (33):** Completion validation flag")
    lines.append("")

    # Extract every row
    lines.append("### Complete Row Data\n")
    lines.append("```")

    for r in range(1, ws.max_row + 1):
        # Collect all non-empty cells
        cells = []
        for c in range(1, max_col + 1):
            v = ws.cell(r, c).value
            if v is not None:
                val_str = str(v).replace("\n", "↵").replace("\r", "")
                # Cap individual cell values at 200 chars
                if len(val_str) > 200:
                    val_str = val_str[:200] + "…"
                col_letter = get_column_letter(c)
                cells.append(f"[{col_letter}{c}]={val_str}")

        if cells:
            line = f"R{r}: {' | '.join(cells)}"
            lines.append(line)

    lines.append("```\n")

    return "\n".join(lines)


# ============================================================================
# SECTION 3: COSTING SHEET — Complete Pricing Engine
# ============================================================================

def extract_costing_sheet(wb):
    lines = []
    lines.append("## SECTION 3: COSTING SHEET — Complete Pricing Engine\n")

    if 'Costing' not in wb.sheetnames:
        lines.append("*No Costing sheet found.*\n")
        return "\n".join(lines)

    ws = wb['Costing']
    lines.append(f"**Total Rows:** {ws.max_row}  |  **Total Cols:** {ws.max_column}\n")

    # Define the key columns for costing
    # A=1, B=2, C=3, D=4, E=5, F=6, G=7, H=8, I=9, J=10, K=11,
    # L=12, M=13, N=14, O=15, P=16, Q=17, R=18, S=19, T=20, U=21, V=22
    # AT=46, AU=47

    key_cols = {
        1: "A_XMarker", 2: "B_ItemName", 3: "C", 4: "D_InputA", 5: "E_InputB",
        6: "F_Qty", 7: "G_UOM", 8: "H_MatsCost", 9: "I_AdjMats",
        10: "J_MatsMarkup", 11: "K_MatsTotal",
        13: "M_LabourDesc", 14: "N_HoursPerUnit", 15: "O_TotalHours",
        16: "P_BaseLabRate", 17: "Q_AdjLabRate", 18: "R_LabMarkup",
        19: "S_LabTotal", 20: "T_LineTotal",
        21: "U_ContractorMats", 22: "V_ContractorHours",
        46: "AT_DataEntered", 47: "AU_NeedsComplete"
    }

    lines.append("### Complete Row Data (All Formula Columns)\n")
    lines.append("```")

    for r in range(1, ws.max_row + 1):
        cells = []
        # Get ALL columns that have data, not just key columns
        for c in range(1, min(ws.max_column + 1, 50)):
            v = ws.cell(r, c).value
            if v is not None:
                val_str = str(v).replace("\n", "↵").replace("\r", "")
                if len(val_str) > 300:
                    val_str = val_str[:300] + "…"
                col_letter = get_column_letter(c)
                col_label = key_cols.get(c, col_letter)
                cells.append(f"[{col_letter}{c}]={val_str}")

        if cells:
            lines.append(f"R{r}: {' | '.join(cells)}")

    lines.append("```\n")

    return "\n".join(lines)


# ============================================================================
# SECTION 4: CF CSV UPLOAD SHEET
# ============================================================================

def extract_cf_csv_sheet(wb):
    lines = []
    lines.append("## SECTION 4: CF CSV UPLOAD SHEET\n")

    if 'CF CSV Upload' not in wb.sheetnames:
        lines.append("*No CF CSV Upload sheet found.*\n")
        return "\n".join(lines)

    ws = wb['CF CSV Upload']
    lines.append(f"**Total Rows:** {ws.max_row}  |  **Total Cols:** {ws.max_column}\n")

    # Headers
    lines.append("### Column Headers (Row 1)\n")
    lines.append("| Col | Letter | Header |")
    lines.append("|-----|--------|--------|")
    for c in range(1, ws.max_column + 1):
        v = cell_val_full(ws, 1, c)
        if v:
            lines.append(f"| {c} | {get_column_letter(c)} | {safe_md(v)} |")
    lines.append("")

    # All data rows with full formulas
    lines.append("### Complete Row Data\n")
    lines.append("```")

    for r in range(1, ws.max_row + 1):
        cells = []
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if v is not None:
                val_str = str(v).replace("\n", "↵").replace("\r", "")
                if len(val_str) > 300:
                    val_str = val_str[:300] + "…"
                col_letter = get_column_letter(c)
                cells.append(f"[{col_letter}{c}]={val_str}")
        if cells:
            lines.append(f"R{r}: {' | '.join(cells)}")

    lines.append("```\n")

    return "\n".join(lines)


# ============================================================================
# SECTION 5: CUSTOMER SUMMARY SHEET
# ============================================================================

def extract_customer_summary(wb):
    lines = []
    lines.append("## SECTION 5: CUSTOMER SUMMARY SHEET\n")

    if 'Customer Summary' not in wb.sheetnames:
        lines.append("*No Customer Summary sheet found.*\n")
        return "\n".join(lines)

    ws = wb['Customer Summary']
    lines.append(f"**Total Rows:** {ws.max_row}  |  **Total Cols:** {ws.max_column}\n")

    lines.append("### Complete Row Data\n")
    lines.append("```")

    for r in range(1, ws.max_row + 1):
        cells = []
        for c in range(1, min(ws.max_column + 1, 20)):
            v = ws.cell(r, c).value
            if v is not None:
                val_str = str(v).replace("\n", "↵").replace("\r", "")
                if len(val_str) > 300:
                    val_str = val_str[:300] + "…"
                col_letter = get_column_letter(c)
                cells.append(f"[{col_letter}{c}]={val_str}")
        if cells:
            lines.append(f"R{r}: {' | '.join(cells)}")

    lines.append("```\n")

    return "\n".join(lines)


# ============================================================================
# SECTION 6: DATA VALIDATION LISTS
# ============================================================================

def extract_data_validations(wb):
    lines = []
    lines.append("## SECTION 6: DATA VALIDATION & DROPDOWN LISTS\n")

    # Part A: Data Lists sheet
    if 'Data Lists' in wb.sheetnames:
        ws = wb['Data Lists']
        lines.append("### 6A: Data Lists Sheet\n")
        lines.append("```")
        for r in range(1, ws.max_row + 1):
            cells = []
            for c in range(1, min(ws.max_column + 1, 10)):
                v = ws.cell(r, c).value
                if v is not None:
                    val_str = str(v).replace("\n", "↵")
                    col_letter = get_column_letter(c)
                    cells.append(f"[{col_letter}{c}]={val_str}")
            if cells:
                lines.append(f"R{r}: {' | '.join(cells)}")
        lines.append("```\n")

    # Part B: Data validation rules on each sheet
    lines.append("### 6B: Data Validation Rules Per Sheet\n")

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if not hasattr(ws, 'data_validations') or ws.data_validations is None:
            continue

        dvs = list(ws.data_validations.dataValidation)
        if not dvs:
            continue

        lines.append(f"#### Sheet: {sheet_name} ({len(dvs)} validation rules)\n")
        lines.append("| # | Cell Range | Type | Operator | Formula1 | Formula2 | Allow Blank | Show List |")
        lines.append("|---|-----------|------|----------|----------|----------|-------------|-----------|")

        for i, dv in enumerate(dvs, 1):
            sqref = str(dv.sqref) if dv.sqref else ""
            dv_type = str(dv.type) if dv.type else ""
            operator = str(dv.operator) if dv.operator else ""
            formula1 = safe_md(str(dv.formula1)) if dv.formula1 else ""
            formula2 = safe_md(str(dv.formula2)) if dv.formula2 else ""
            allow_blank = str(dv.allow_blank) if dv.allow_blank is not None else ""
            show_list = str(dv.showDropDown) if hasattr(dv, 'showDropDown') else ""

            # Cap formula length
            if len(formula1) > 200:
                formula1 = formula1[:200] + "…"

            lines.append(f"| {i} | {sqref} | {dv_type} | {operator} | {formula1} | {formula2} | {allow_blank} | {show_list} |")

        lines.append("")

    return "\n".join(lines)


# ============================================================================
# SECTION 7: VBA CODE
# ============================================================================

def extract_vba(filepath):
    lines = []
    lines.append("## SECTION 7: VBA CODE\n")

    if not HAS_OLETOOLS:
        lines.append("*oletools not installed — VBA extraction skipped.*\n")
        lines.append("Install with: `pip install oletools`\n")
        return "\n".join(lines)

    try:
        vba = VBA_Parser(filepath)
        if not vba.detect_vba_macros():
            lines.append("*No VBA macros detected.*\n")
            return "\n".join(lines)

        macro_count = 0
        for (filename, stream_path, vba_filename, vba_code) in vba.extract_macros():
            if not vba_code.strip():
                continue
            macro_count += 1
            lines.append(f"### Module: {vba_filename}\n")
            lines.append(f"**Stream:** `{stream_path}`\n")
            lines.append("```vb")
            lines.append(vba_code)
            lines.append("```\n")

        if macro_count == 0:
            lines.append("*VBA modules found but all were empty.*\n")

        vba.close()
    except Exception as e:
        lines.append(f"*VBA extraction failed: {e}*\n")

    return "\n".join(lines)


# ============================================================================
# SECTION 8: CROSS-SHEET REFERENCES
# ============================================================================

def extract_cross_sheet_refs(wb):
    lines = []
    lines.append("## SECTION 8: CROSS-SHEET REFERENCES\n")

    # Pattern to find sheet references in formulas
    # Matches: SheetName!CellRef, 'Sheet Name'!CellRef, [1]SheetName!CellRef
    ref_pattern = re.compile(r"""
        (?:\[[\d]+\])?          # Optional external workbook ref [1]
        (?:'([^']+)'|(\w+))     # Sheet name (quoted or unquoted)
        !                       # Separator
        (\$?[A-Z]+\$?\d+       # Cell reference
        (?::\$?[A-Z]+\$?\d+)?) # Optional range
    """, re.VERBOSE)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        refs_found = {}

        for r in range(1, ws.max_row + 1):
            for c in range(1, min(ws.max_column + 1, 50)):
                v = ws.cell(r, c).value
                if v and isinstance(v, str) and "!" in v:
                    matches = ref_pattern.findall(str(v))
                    for match in matches:
                        target_sheet = match[0] or match[1]
                        target_cell = match[2]
                        source = f"{sheet_name}!{get_column_letter(c)}{r}"
                        key = f"{target_sheet}!{target_cell}"
                        if key not in refs_found:
                            refs_found[key] = []
                        refs_found[key].append(source)

        if refs_found:
            lines.append(f"### References FROM: {sheet_name}\n")
            lines.append("| Target | Referenced By |")
            lines.append("|--------|-------------|")
            for target, sources in sorted(refs_found.items()):
                src_list = ", ".join(sources[:10])
                if len(sources) > 10:
                    src_list += f" (+{len(sources)-10} more)"
                lines.append(f"| {target} | {src_list} |")
            lines.append("")

    return "\n".join(lines)


# ============================================================================
# SECTION 9: NAMED RANGES
# ============================================================================

def extract_named_ranges(wb):
    lines = []
    lines.append("## SECTION 9: NAMED RANGES\n")

    # Handle both old and new openpyxl API
    try:
        names = list(wb.defined_names.definedNameList)
    except AttributeError:
        try:
            names = list(wb.defined_names.values())
        except Exception:
            names = []
            for name in wb.defined_names:
                names.append(wb.defined_names[name])

    if not names:
        lines.append("*No named ranges defined.*\n")
        return "\n".join(lines)

    lines.append(f"**Count:** {len(names)}\n")
    lines.append("| Name | Value | Scope |")
    lines.append("|------|-------|-------|")

    for dn in names:
        name = dn.name if hasattr(dn, 'name') else str(dn)
        value = safe_md(str(dn.attr_text)) if hasattr(dn, 'attr_text') else safe_md(str(dn.value)) if hasattr(dn, 'value') else ""
        scope = "Workbook" if (not hasattr(dn, 'localSheetId') or dn.localSheetId is None) else f"Sheet #{dn.localSheetId}"
        lines.append(f"| {name} | {value} | {scope} |")

    lines.append("")
    return "\n".join(lines)


# ============================================================================
# SECTION 10: CONDITIONAL FORMATTING
# ============================================================================

def extract_conditional_formatting(wb):
    lines = []
    lines.append("## SECTION 10: CONDITIONAL FORMATTING\n")

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if not hasattr(ws, 'conditional_formatting') or not ws.conditional_formatting:
            continue

        cf_list = list(ws.conditional_formatting)
        if not cf_list:
            continue

        lines.append(f"### Sheet: {sheet_name} ({len(cf_list)} rules)\n")

        for i, cf in enumerate(cf_list, 1):
            cells = str(cf)
            lines.append(f"**Rule {i}** — Range: `{cells}`\n")
            for rule in cf.rules:
                lines.append(f"- Type: `{rule.type}`")
                if hasattr(rule, 'priority'):
                    lines.append(f"  Priority: {rule.priority}")
                if hasattr(rule, 'operator') and rule.operator:
                    lines.append(f"  Operator: {rule.operator}")
                if hasattr(rule, 'formula') and rule.formula:
                    lines.append(f"  Formula: `{rule.formula}`")
                if hasattr(rule, 'dxf') and rule.dxf:
                    dxf = rule.dxf
                    if dxf.font:
                        lines.append(f"  Font: color={dxf.font.color}, bold={dxf.font.bold}")
                    if dxf.fill:
                        lines.append(f"  Fill: {dxf.fill}")
                    if dxf.border:
                        lines.append(f"  Border: {dxf.border}")
                lines.append("")
        lines.append("")

    return "\n".join(lines)


# ============================================================================
# SECTION 11: ALL OTHER SHEETS (catch-all)
# ============================================================================

def extract_other_sheets(wb):
    """Extract content from sheets not already covered."""
    covered = {'Report', 'Costing', 'CF CSV Upload', 'Customer Summary', 'Data Lists'}
    lines = []
    lines.append("## SECTION 11: OTHER SHEETS\n")

    for sheet_name in wb.sheetnames:
        if sheet_name in covered:
            continue

        ws = wb[sheet_name]
        if ws.max_row == 0 or ws.max_row is None:
            continue

        lines.append(f"### Sheet: {sheet_name}\n")
        lines.append(f"**Rows:** {ws.max_row}  |  **Cols:** {ws.max_column}\n")
        lines.append("```")

        for r in range(1, ws.max_row + 1):
            cells = []
            for c in range(1, min(ws.max_column + 1, 40)):
                v = ws.cell(r, c).value
                if v is not None:
                    val_str = str(v).replace("\n", "↵").replace("\r", "")
                    if len(val_str) > 200:
                        val_str = val_str[:200] + "…"
                    col_letter = get_column_letter(c)
                    cells.append(f"[{col_letter}{c}]={val_str}")
            if cells:
                lines.append(f"R{r}: {' | '.join(cells)}")

        lines.append("```\n")

    return "\n".join(lines)


# ============================================================================
# MAIN
# ============================================================================

def main():
    if len(sys.argv) < 2:
        print("Usage: python extract_workbook.py <path_to_xlsm> [output_dir]")
        print("")
        print("Example:")
        print("  python extract_workbook.py Copy_of_Damp_Costing_v48_CF_-_220126.xlsm ./output")
        sys.exit(1)

    filepath = sys.argv[1]
    output_dir = sys.argv[2] if len(sys.argv) > 2 else "."

    if not os.path.exists(filepath):
        print(f"ERROR: File not found: {filepath}")
        sys.exit(1)

    os.makedirs(output_dir, exist_ok=True)

    print(f"Opening workbook: {filepath}")
    print("  (Using data_only=False to capture formulas)")

    # Load with formulas (not calculated values)
    wb = openpyxl.load_workbook(filepath, read_only=False, data_only=False, keep_vba=False)

    survey_type = detect_survey_type(wb)
    print(f"  Detected survey type: {survey_type}")
    print(f"  Sheets: {wb.sheetnames}")

    output_file = os.path.join(output_dir, f"{survey_type}_WORKBOOK_ANALYSIS.md")
    print(f"  Output: {output_file}")
    print("")

    sections = []

    # Header
    sections.append(f"# {survey_type} Workbook — Forensic Analysis\n")
    sections.append(f"**Source:** `{os.path.basename(filepath)}`  ")
    sections.append(f"**Extracted:** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}  ")
    sections.append(f"**Tool:** extract_workbook.py v1.0\n")
    sections.append("---\n")

    # Section 1: Overview
    print("Extracting Section 1: Workbook Overview...")
    sections.append(extract_overview(wb, filepath))

    # Section 2: Report Sheet
    print("Extracting Section 2: Report Sheet (this may take a moment)...")
    sections.append(extract_report_sheet(wb))

    # Section 3: Costing Sheet
    print("Extracting Section 3: Costing Sheet...")
    sections.append(extract_costing_sheet(wb))

    # Section 4: CF CSV Upload
    print("Extracting Section 4: CF CSV Upload Sheet...")
    sections.append(extract_cf_csv_sheet(wb))

    # Section 5: Customer Summary
    print("Extracting Section 5: Customer Summary...")
    sections.append(extract_customer_summary(wb))

    # Section 6: Data Validation
    print("Extracting Section 6: Data Validation Rules...")
    sections.append(extract_data_validations(wb))

    wb.close()

    # Section 7: VBA (needs separate parser)
    print("Extracting Section 7: VBA Code...")
    sections.append(extract_vba(filepath))

    # Re-open for remaining sections
    wb = openpyxl.load_workbook(filepath, read_only=False, data_only=False, keep_vba=False)

    # Section 8: Cross-sheet refs
    print("Extracting Section 8: Cross-Sheet References...")
    sections.append(extract_cross_sheet_refs(wb))

    # Section 9: Named Ranges
    print("Extracting Section 9: Named Ranges...")
    sections.append(extract_named_ranges(wb))

    # Section 10: Conditional Formatting
    print("Extracting Section 10: Conditional Formatting...")
    sections.append(extract_conditional_formatting(wb))

    # Section 11: All other sheets
    print("Extracting Section 11: Other Sheets...")
    sections.append(extract_other_sheets(wb))

    wb.close()

    # Write output
    print(f"\nWriting {output_file}...")
    content = "\n".join(sections)
    with open(output_file, "w", encoding="utf-8") as f:
        f.write(content)

    file_size = os.path.getsize(output_file)
    print(f"  Done! File size: {file_size:,} bytes ({file_size/1024:.1f} KB)")
    print(f"\n  Output: {output_file}")


if __name__ == "__main__":
    main()

