#!/usr/bin/env python3
"""Generate oracle cellmaps from the ORIGINAL workbooks.

Reads a Costing sheet directly from the .xlsm, finds every priced line row
(K = F*I*(1+J) signature), section-total rows ("Section Price Adjustment %"),
the totals block, travel inputs, and rates row — and emits a cellmap JSON
compatible with run_oracle.py. Template line_keys are matched with the same
logic (and MANUAL_MAP) as the rates audit, so the differ can join lines.

Usage (from repo root):
    python3 parity/oracle/build_cellmap.py condensation timber woodworm
"""
import importlib.util
import json
import re
import sys
from pathlib import Path

REPO = Path(__file__).resolve().parents[2]
CELLMAPS = REPO / "parity" / "oracle" / "cellmaps"

# Reuse extraction + matching from the rates audit (single source of truth)
spec = importlib.util.spec_from_file_location("rates_audit", REPO / "parity" / "audit" / "rates_audit.py")
ra = importlib.util.module_from_spec(spec)
spec.loader.exec_module(ra)

COL_LETTER = {v: k for k, v in ra.COL.items()}
LABEL_COLS = {3: "C", 5: "E", 8: "H", 10: "J", 13: "M", 21: "U"}

WORKBOOK_IDS = {
    "damp": "damp_v48",
    "condensation": "condensation_v37",
    "timber": "timber_v33",
    "woodworm": "woodworm_v26",
}


def slug(s, maxlen=40):
    s = re.sub(r"[^a-z0-9]+", "_", str(s).lower()).strip("_")
    return s[:maxlen].rstrip("_") or "row"


def full_grid(path):
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=False, read_only=True)
    sheet = next(s for s in wb.sheetnames if s.strip().lower() == "costing")
    ws = wb[sheet]
    grid = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                grid[(cell.row, cell.column)] = cell.value
    wb.close()
    return grid


def find_label_cell(grid, contains, col=None, exact=False):
    """Find (row, col) whose string value matches; returns list."""
    out = []
    for (r, c), v in grid.items():
        if col and c != col:
            continue
        s = str(v).strip().lower()
        t = contains.lower()
        if (s == t) if exact else (t in s):
            out.append((r, c))
    return sorted(out)


def build(survey_type):
    fname = dict(ra.WORKBOOKS)[survey_type]
    path = ra.WB_DIR / fname
    print(f"== {survey_type}: {fname}")
    lines, rates = ra.extract_workbook(path)
    grid = full_grid(path)
    templates, materials, config = ra.build_template_index()
    matches, used = ra.match_templates(lines, templates, survey_type)

    # --- lines
    out_lines, unmatched = {}, []
    for ln in lines:
        ts = matches.get(ln["row"])
        keys = [t["line_key"] for t in ts] if ts else None
        name = f"{slug(ln['label'])}_r{ln['row']}"
        entry = {
            "row": ln["row"],
            "input": "none" if ln["input"] == "derived" else ln["input"],
            "label": ln["label"],
            "line_keys": keys,
            "section": None,
        }
        if ln["input"] == "derived":
            entry["derived"] = ln["derived_formula"]
        if ln["input"] == "de":
            entry["d_label"], entry["e_label"] = "d", "e"
        if ra.is_zero_feeder(ln):
            entry["zero_feeder"] = True
        out_lines[name] = entry
        if not keys and not ra.is_zero_feeder(ln):
            unmatched.append(f"R{ln['row']} {ln['label'][:50]}")

    # --- section totals + nearest preceding header for naming
    header_rows = {}
    for (r, c), v in grid.items():
        if c == 13 and isinstance(v, str) and re.match(r"^=B\d+$", v.replace(" ", "")):
            b = grid.get((r, 2))
            if b:
                header_rows[r] = str(b).strip()
    sections = {}
    for (r, c) in find_label_cell(grid, "section price adjustment", col=2):
        hdr = max((hr for hr in header_rows if hr < r), default=None)
        name = f"{slug(header_rows.get(hdr, ''), 30)}_r{r}" if hdr else f"section_r{r}"
        sections[name] = {"total_row": r, "adjust_cell": f"F{r}"}
        # annotate lines belonging to this section (rows between prev total and r)
    prev = 0
    for name in sorted(sections, key=lambda n: sections[n]["total_row"]):
        r = sections[name]["total_row"]
        for ln_name, e in out_lines.items():
            if prev < e["row"] < r and e["section"] is None:
                e["section"] = name
        prev = r

    # --- totals block by label search
    totals, notes = {}, []

    def put(key, cells):
        if cells:
            totals[key] = cells[0]
        else:
            notes.append(f"MISSING totals label for {key}")

    def kcell(hits):
        return [f"K{r}" for r, _ in hits]

    put("materials_subtotal", kcell(find_label_cell(grid, "materials sub tot", col=5)))
    lab_hits = find_label_cell(grid, "labour sub tot", col=5)
    put("labour_subtotal", kcell(lab_hits))
    put("labour_hours_subtotal", [f"O{r}" for r, _ in find_label_cell(grid, "labour hours sub total", col=13)])
    trav_h = find_label_cell(grid, "travel", col=8, exact=True)
    put("travel_price", kcell(trav_h))
    put("travel_hours", [f"O{r}" for r, _ in find_label_cell(grid, "hours travel", col=13)])
    put("subtotal_ex_vat", kcell(find_label_cell(grid, "price", col=8, exact=True)))
    put("total_hours", [f"O{r}" for r, _ in find_label_cell(grid, "total hours", col=13)])
    put("vat", kcell(find_label_cell(grid, "vat", col=8, exact=True)))
    put("total_inc_vat", kcell(find_label_cell(grid, "total price inc vat", col=8)))
    put("days", kcell(find_label_cell(grid, "no of days", col=10)))
    if lab_hits:
        r = lab_hits[0][0]
        totals["contractor_materials_total"] = f"U{r}"
        totals["contractor_pay_total"] = f"V{r}"
    ct = find_label_cell(grid, "travel", col=21, exact=True)
    if ct:
        totals["contractor_travel"] = f"V{ct[0][0]}"
    tot = find_label_cell(grid, "tot", col=21, exact=True)
    if tot:
        totals["contractor_total"] = f"V{tot[0][0]}"

    # --- travel inputs
    travel_inputs = {}
    d_hits = find_label_cell(grid, "from office", col=10)
    m_hits = find_label_cell(grid, "no of men travelling", col=10)
    if d_hits:
        travel_inputs["distance_one_way_miles"] = f"K{d_hits[0][0]}"
    if m_hits:
        travel_inputs["men_travelling"] = f"K{m_hits[0][0]}"
    if len(travel_inputs) < 2:
        notes.append("MISSING travel input cells")

    lr = rates.get("labour_row")
    rates_out = {"labour_rate": f"D{lr}", "contractor_rate": f"E{lr}", "vehicle_cost_per_mile": f"J{lr}"} if lr else {}

    cellmap = {
        "workbook_file": f"workbook_extraction/workbooks/{fname}",
        "workbook_id": WORKBOOK_IDS[survey_type],
        "sheet": "COSTING",
        "comment": f"AUTO-GENERATED by build_cellmap.py from the live workbook ({survey_type}). "
                   "Lines matched to costing_line_templates via rates-audit logic. "
                   "Verify totals-block cells before first trusted run.",
        "line_columns": {"quantity": "F", "unit_material_cost": "H", "materials": "K", "hours": "O",
                          "labour": "S", "total": "T", "contractor_materials": "U", "contractor_pay": "V"},
        "lines": out_lines,
        "sections": sections,
        "travel_inputs": travel_inputs,
        "totals": totals,
        "rates": rates_out,
    }
    out_path = CELLMAPS / f"{WORKBOOK_IDS[survey_type]}.json"
    out_path.write_text(json.dumps(cellmap, indent=2))
    print(f"  {len(out_lines)} lines ({len(unmatched)} unmatched), {len(sections)} sections, "
          f"totals={len(totals)}, -> {out_path.relative_to(REPO)}")
    for u in unmatched:
        print(f"    UNMATCHED: {u}")
    for n in notes:
        print(f"    NOTE: {n}")


if __name__ == "__main__":
    for st in (sys.argv[1:] or ["condensation", "timber", "woodworm"]):
        build(st)
