#!/usr/bin/env python3
"""DATA-TRACK RATES AUDIT — workbooks vs costing_line_templates.

Parses every costing line row from the four ORIGINAL workbooks (formulas read
directly from the .xlsm via openpyxl — not from any prior analysis), derives
each line's canonical rate tuple, then field-diffs the platform's template
data following the ENGINE'S OWN value precedence:

    effective_unit_cost = (materials_catalog[product_key] / coverage_rate
                           if product_key in catalog
                           else params.cost_per_coverage_unit
                           else base_unit_cost) * wastage_factor

Outputs (parity/audit/):
    RATES_AUDIT.md            human report, classified findings
    proposed-corrections.sql  DRAFT corrections — DO NOT APPLY until the
                              harness-gated fix batch (costing freeze)
    raw/<workbook>.json       extracted line evidence per workbook

Run from repo root: python3 parity/audit/rates_audit.py
"""
import json
import re
import sys
from difflib import SequenceMatcher
from pathlib import Path

import openpyxl

REPO = Path(__file__).resolve().parents[2]
AUDIT = REPO / "parity" / "audit"
FIXTURES = REPO / "parity" / "fixtures"
WB_DIR = REPO / "workbook_extraction" / "workbooks"

WORKBOOKS = [
    ("damp", "Copy of Damp Costing v48 CF - 220126.xlsm"),
    ("condensation", "Copy of Condensation PIV Costing v37 CF - 131125.xlsm"),
    ("timber", "Copy of Timber Costing v33 - CF - 131125.xlsm"),
    ("woodworm", "Copy of Woodworm Costing v26 - CF Version - 220126.xlsm"),
]

COL = {"B": 2, "D": 4, "E": 5, "F": 6, "G": 7, "H": 8, "I": 9, "J": 10,
       "K": 11, "N": 14, "O": 15, "P": 16, "R": 18, "S": 19, "T": 20}

TOL = 5e-7          # exact match tolerance
TRUNC_TOL = 5e-3    # matches to ~3dp -> classified TRUNCATED, not WRONG

ARITH = re.compile(r"^[0-9+\-*/(). ]+$")
PACK_RE = re.compile(
    r"^=?IF\(F\d+=0,0,\(_?x?l?f?n?\.?CEILING\.MATH\(F\d+,([0-9.]+)\)"
    r"\*\(\(([0-9./]+)\)\*([0-9.]+)\)\)/F\d+\)$"
)


def norm_formula(f):
    return str(f).replace(" ", "").replace("_xlfn.", "") if f is not None else ""


def safe_eval(expr):
    expr = str(expr).strip().lstrip("=")
    if ARITH.match(expr):
        try:
            return float(eval(expr, {"__builtins__": {}}, {}))
        except Exception:
            return None
    return None


def classify_H(raw):
    """Return (model, payload) for a material-cost cell."""
    if raw is None or raw == "":
        return ("empty", None)
    if isinstance(raw, (int, float)):
        return ("const", float(raw))
    s = norm_formula(raw)
    v = safe_eval(s)
    if v is not None:
        return ("const", v)
    m = re.match(
        r"^=IF\(F\d+=0,0,\(CEILING\.MATH\(F\d+,([0-9.]+)\)\*\(\(([0-9.]+)/([0-9.]+)\)\*([0-9.]+)\)\)/F\d+\)$",
        s,
    )
    if m:
        step, price, div, factor = (float(m.group(i)) for i in (1, 2, 3, 4))
        return ("pack", {"step": step, "pack_price": price, "div": div,
                         "factor": factor, "unit_cost": price / div * factor})
    # guarded constant, e.g. digital DPC: =IF(OR(...),0,650)
    m = re.match(r"^=IF\(.+,0,([0-9.]+)\)$", s)
    if m:
        return ("const", float(m.group(1)))
    return ("special", s)


def classify_N(raw, o_raw, row):
    """Labour model: hours per unit, or block/bespoke."""
    o = norm_formula(o_raw)
    n_val = safe_eval(raw) if not isinstance(raw, (int, float)) else float(raw)
    m = re.match(rf"^=ROUNDUP\(F{row}/([0-9.]+),0\)\*([0-9.]+)$", o)
    if m:
        return ("block", {"block_size": float(m.group(1)), "hours_per_block": float(m.group(2))})
    if o and o not in (f"=F{row}*N{row}",):
        return ("special_o", {"o_formula": o, "n_value": n_val})
    if n_val is not None:
        return ("per_unit", n_val)
    return ("special_n", {"n_formula": norm_formula(raw)})


def extract_workbook(path):
    wb = openpyxl.load_workbook(path, data_only=False, read_only=True)
    sheet_name = next((s for s in wb.sheetnames if s.strip().lower() == "costing"), None)
    if not sheet_name:
        raise RuntimeError(f"no Costing sheet in {path.name}")
    ws = wb[sheet_name]
    grid = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None and cell.column in COL.values():
                grid[(cell.row, cell.column)] = cell.value
    wb.close()

    lines, rates = [], {}
    max_row = max((r for r, _ in grid), default=0)
    for r in range(1, max_row + 1):
        k = norm_formula(grid.get((r, COL["K"])))
        std_sig = re.match(rf"^=F{r}\*I{r}\*\(1\+J{r}\)$", k)
        # Timber resin rows use whole-pack pricing: =((ROUNDUP(F/30,0))*I)*(1+J59)
        pack_sig = re.match(rf"^=\(\(ROUNDUP\(F{r}/([0-9.]+),0\)\)\*I{r}\)\*\(1\+J\d+\)$", k)
        if not (std_sig or pack_sig):
            continue
        k_pack_size = float(pack_sig.group(1)) if pack_sig else None
        f_raw = grid.get((r, COL["F"]))
        f_norm = norm_formula(f_raw) if isinstance(f_raw, str) else ""
        input_kind = ("de" if f_norm == f"=D{r}*E{r}"
                      else "derived" if f_norm.startswith("=")
                      else "f")
        h_model, h = classify_H(grid.get((r, COL["H"])))
        n_model, n = classify_N(grid.get((r, COL["N"])), grid.get((r, COL["O"])), r)
        p = norm_formula(grid.get((r, COL["P"])))
        pm = re.search(r"\$?D\$?(\d+)", p)
        if pm:
            rates.setdefault("labour_row", int(pm.group(1)))
        lines.append({
            "row": r,
            "label": str(grid.get((r, COL["B"]), "")).strip(),
            "uom": str(grid.get((r, COL["G"]), "")).strip(),
            "input": input_kind,
            "derived_formula": f_norm if input_kind == "derived" else None,
            "h_model": h_model, "h": h,
            "n_model": n_model, "n": n,
            "k_pack_size": k_pack_size,
            "material_markup": safe_eval(grid.get((r, COL["J"]))) if grid.get((r, COL["J"])) is not None else None,
            "labour_markup": safe_eval(grid.get((r, COL["R"]))) if grid.get((r, COL["R"])) is not None else None,
        })
    lr = rates.get("labour_row")
    if lr:
        rates["hourly_labour_rate"] = safe_eval(grid.get((lr, COL["D"])))
        rates["contractor_hourly_rate"] = safe_eval(grid.get((lr, COL["E"])))
        rates["vehicle_cost_per_mile"] = safe_eval(grid.get((lr, COL["J"])))
    return lines, rates


def norm_label(s):
    return re.sub(r"[^a-z0-9]+", " ", str(s).lower()).strip()


# Explicit (survey_type, workbook_row) -> line_key(s) where description drift or
# workbook structure defeats text matching. Rows 46-48 (2m membrane feeders)
# carry no pricing (H=0, N=0) — the priced row is the R49 subtotal, which maps
# to the platform's single wall_membrane_2m line.
MANUAL_MAP = {
    ("damp", 42): ["dpc_installation_digital"],
    ("damp", 49): ["wall_membrane_2m"],
    ("damp", 50): ["membrane_plugs"],
    ("damp", 57): ["difficulty_hours_walls"],
    ("damp", 64): ["wall_floor_fillet_tanking"],
    ("damp", 65): ["difficulty_hours_tanking"],
    ("damp", 71): ["wall_floor_fillet_resin"],
    ("damp", 73): ["difficulty_hours_resin"],
    ("damp", 79): ["warmline_iwi", "warmline_iwi_adhesive"],
    ("damp", 82): ["thin_coat_angle_2_4m"],
    ("damp", 83): ["thin_coat_angle_3m"],
    ("damp", 84): ["difficulty_hours_plastering"],
    ("damp", 99): ["difficulty_hours_joists"],
    ("damp", 101): ["weyrock_flooring_18mm"],
    ("damp", 102): ["weyrock_flooring_22mm"],
    ("damp", 103): ["std_tongue_groove_floorboards"],
    ("damp", 104): ["victorian_tongue_groove_floorboards"],
    ("damp", 105): ["engineered_flooring_sheet"],
    ("damp", 106): ["structural_engineered_flooring_sheet"],
    ("damp", 108): ["difficulty_hours_decking"],
    ("damp", 119): ["difficulty_hours_spray"],
    ("timber", 73): ["warmline_internal_wall_insulation", "warmline_iwi_adhesive"],
    # Condensation: four identical "Diamond core 107mm hole" labels and three
    # identical electrical-pack labels — pin each row to its section's template
    # (matcher otherwise assigns them in encounter order).
    ("condensation", 23): ["electrical_pack_fused_spur_cable_jb"],
    ("condensation", 29): ["electrical_pack_fused_spur_cable_jb_piv_wall"],
    ("condensation", 46): ["electrical_pack_fused_spur_cable_jb_humidistat_fan"],
    ("condensation", 41): ["diamond_core_107mm_hole"],
    ("condensation", 48): ["diamond_core_107mm_hole_humidistat_fan"],
    ("condensation", 53): ["diamond_core_107mm_hole_passive_vent"],
    ("condensation", 58): ["diamond_core_107mm_hole_dryaire_cvent"],
}


def is_zero_feeder(ln):
    """2m-membrane feeder rows: quantity contributors with no pricing content."""
    return (ln["h_model"] == "const" and not ln["h"]
            and ln["n_model"] == "per_unit" and not ln["n"])


def build_template_index():
    templates = json.loads((FIXTURES / "costing_line_templates.json").read_text())
    sections = {s["id"]: s for s in json.loads((FIXTURES / "costing_sections.json").read_text())}
    for t in templates:
        sec = sections.get(t["section_id"], {})
        t["survey_type"] = sec.get("survey_type")
        t["section_key"] = sec.get("section_key")
    materials = {m["product_key"]: m["unit_cost"]
                 for m in json.loads((FIXTURES / "materials_catalog.json").read_text())
                 if m.get("is_active")}
    config = {c["config_key"]: c["config_value"]
              for c in json.loads((FIXTURES / "pricing_config.json").read_text())}
    return templates, materials, config


def engine_effective_material(t, materials, config):
    """Mirror calcStandard/calcCeilingCoverage value precedence."""
    params = t.get("formula_params") or {}
    wastage = t["wastage_factor"] if t["wastage_factor"] is not None else config.get("default_wastage_factor", 1.1)
    if t["cost_formula"] == "whole_pack":
        size = params.get("pack_size") or t["coverage_rate"] or 1
        cost = params.get("pack_cost") or t["base_unit_cost"] or 0
        return cost / size * wastage, {"source": "whole_pack", "pack_size": size, "pack_cost": cost, "wastage": wastage}
    if t["cost_formula"] == "ceiling_coverage":
        coverage = params.get("coverage_rate", t["coverage_rate"])
        pk = params.get("product_key")
        if pk and pk in materials:
            unit = materials[pk] / coverage
            src = f"catalog[{pk}]={materials[pk]}/{coverage}"
        elif params.get("cost_per_coverage_unit") is not None:
            unit = params["cost_per_coverage_unit"]
            src = "params.cost_per_coverage_unit"
        else:
            unit = t["base_unit_cost"] or 0
            src = "base_unit_cost"
        return unit * wastage, {"coverage": coverage, "source": src, "wastage": wastage}
    unit = t["base_unit_cost"]
    if unit in (None, 0) and params.get("product_key") in materials:
        unit = materials[params["product_key"]]
    return (unit or 0) * wastage, {"source": "base_unit_cost", "wastage": wastage}


def match_templates(lines, templates, survey_type):
    pool = [t for t in templates if t["survey_type"] in (survey_type, "site_preparation")]
    by_desc, by_key = {}, {}
    for t in pool:
        by_desc.setdefault(norm_label(t["description"]), []).append(t)
        by_key.setdefault(t["line_key"], []).append(t)
    matches, used = {}, set()
    # manual overrides first (may map one row to MULTIPLE templates)
    for ln in lines:
        manual = MANUAL_MAP.get((survey_type, ln["row"]))
        if manual:
            ts = [t for k in manual for t in by_key.get(k, [])]
            if ts:
                matches[ln["row"]] = ts
                used.update(t["id"] for t in ts)
    for ln in lines:
        if ln["row"] in matches or is_zero_feeder(ln):
            continue
        lbl = norm_label(ln["label"])
        cands = by_desc.get(lbl, [])
        if not cands:
            scored = sorted(
                ((SequenceMatcher(None, lbl, norm_label(t["description"])).ratio(), t) for t in pool),
                key=lambda x: -x[0],
            )
            cands = [t for r, t in scored[:1] if r >= 0.82]
        free = [t for t in cands if t["id"] not in used]
        if free:
            matches[ln["row"]] = [free[0]]
            used.add(free[0]["id"])
    return matches, used


def diff_line(ln, t, materials, config, findings):
    key = f"{t['survey_type']}/{t['section_key']}/{t['line_key']}"

    def add(cls, field, wb_v, pl_v, note=""):
        findings.append({
            "workbook_row": ln["row"], "label": ln["label"], "template": key,
            "template_id": t["id"], "class": cls, "field": field,
            "workbook": wb_v, "platform": pl_v, "note": note,
        })

    # --- markups
    if ln["material_markup"] is not None and abs(ln["material_markup"] - (t["material_markup"] or 0)) > TOL:
        add("MARKUP", "material_markup", ln["material_markup"], t["material_markup"])
    if ln["labour_markup"] is not None and abs(ln["labour_markup"] - (t["labour_markup"] or 0)) > TOL:
        add("MARKUP", "labour_markup", ln["labour_markup"], t["labour_markup"])

    # --- materials
    eff, meta = engine_effective_material(t, materials, config)
    if t["cost_formula"] == "whole_pack" and ln["h_model"] == "const":
        params2 = t.get("formula_params") or {}
        if abs((params2.get("pack_cost") or 0) - ln["h"]) > TOL:
            add("WRONG_RATE", "pack_cost", ln["h"], params2.get("pack_cost"), "whole_pack")
        if ln.get("k_pack_size") is not None and abs((params2.get("pack_size") or 0) - ln["k_pack_size"]) > TOL:
            add("WRONG_RATE", "pack_size", ln["k_pack_size"], params2.get("pack_size"), "whole_pack")
    elif ln["h_model"] == "const":
        delta = eff - ln["h"]
        if abs(delta) > TOL:
            ratio = eff / ln["h"] if ln["h"] else None
            if ratio and abs(ratio - 1.1) < 1e-3:
                add("EXTRA_WASTAGE", "material_unit_cost", ln["h"], eff,
                    f"platform = workbook x1.1 (wastage_factor={meta['wastage']})")
            elif abs(delta) <= TRUNC_TOL * max(1.0, abs(ln["h"])):
                add("TRUNCATED", "material_unit_cost", ln["h"], eff, meta["source"])
            else:
                add("WRONG_RATE", "material_unit_cost", ln["h"], eff, meta["source"])
    elif ln["h_model"] == "pack":
        wb_unit = ln["h"]["unit_cost"]
        if t["cost_formula"] != "ceiling_coverage":
            add("MODEL_MISMATCH", "cost_formula", "pack/CEILING", t["cost_formula"],
                f"workbook step={ln['h']['step']}")
        else:
            cov = meta.get("coverage")
            if cov is not None and abs(cov - ln["h"]["step"]) > TOL:
                add("WRONG_COVERAGE", "coverage_rate", ln["h"]["step"], cov,
                    f"workbook price divisor={ln['h']['div']}")
            if ln["h"]["div"] != ln["h"]["step"]:
                add("QUIRK", "pack_div_vs_step", ln["h"]["step"], ln["h"]["div"],
                    "workbook CEILING step differs from price divisor — verify representation")
        delta = eff - wb_unit
        if abs(delta) > TOL:
            ratio = eff / wb_unit if wb_unit else None
            if ratio and abs(ratio - 1 / 1.1) < 2e-3:
                add("MISSING_WASTAGE", "material_unit_cost", wb_unit, eff,
                    f"platform = workbook /1.1 (wastage_factor={meta['wastage']}, {meta['source']})")
            elif abs(delta) <= TRUNC_TOL * max(1.0, abs(wb_unit)):
                add("TRUNCATED", "material_unit_cost", wb_unit, eff, meta["source"])
            else:
                add("WRONG_RATE", "material_unit_cost", wb_unit, eff, meta["source"])
    elif ln["h_model"] == "special":
        add("SPECIAL", "material_formula", ln["h"], t["cost_formula"],
            "bespoke workbook formula — needs manual/code-track verification")

    # --- labour
    params = t.get("formula_params") or {}
    if ln["n_model"] == "per_unit":
        pl = t["labour_rate_per_unit"]
        if pl is not None and abs(pl - ln["n"]) > TOL:
            cls = "TRUNCATED" if abs(pl - ln["n"]) <= TRUNC_TOL else "WRONG_RATE"
            add(cls, "labour_rate_per_unit", ln["n"], pl)
    elif ln["n_model"] == "block":
        if (params.get("labour_block_size") != ln["n"]["block_size"]
                or params.get("labour_hours_per_block") != ln["n"]["hours_per_block"]):
            add("WRONG_RATE", "labour_block",
                f"{ln['n']['hours_per_block']}h/{ln['n']['block_size']}",
                f"{params.get('labour_hours_per_block')}h/{params.get('labour_block_size')}")
    elif ln["n_model"] in ("special_o", "special_n"):
        add("SPECIAL", "labour_formula", ln["n"], t["labour_rate_per_unit"],
            "bespoke workbook labour — needs manual/code-track verification")


def generate_corrections(pairs, materials):
    """Emit SQL setting each mismatched value at its TRUE source so that
    engine-effective == workbook exactly:
      - wastage_factor -> the workbook's factor for this line (1.1 pack lines,
        1.0 const lines: the workbook has no universal wastage concept)
      - pack unit precision -> materials_catalog.unit_cost (price*coverage/div),
        params.cost_per_coverage_unit (price/div), or base_unit_cost
      - labour_rate_per_unit -> exact workbook hours/unit
    Skips SPECIAL/bespoke and multi-template rows (code track). Idempotent:
    emits only actual differences.
    """
    tmpl_updates = {}   # template_id -> {col: value}
    param_updates = {}  # template_id -> cost_per_coverage_unit
    catalog_targets = {}  # product_key -> (target, [labels])
    comments = {}

    def want(tid, col, val, label):
        tmpl_updates.setdefault(tid, {})[col] = val
        comments.setdefault(tid, label)

    for survey_type, ln, t in pairs:
        if t["cost_formula"] == "whole_pack":
            continue
        tid = t["id"]
        label = f"{survey_type} R{ln['row']} {ln['label'][:50]}"
        params = t.get("formula_params") or {}
        wastage = t["wastage_factor"] if t["wastage_factor"] is not None else 1.1
        if ln["h_model"] == "const":
            if abs(wastage - 1.0) > 1e-7:
                want(tid, "wastage_factor", 1.0, label)
            base = t["base_unit_cost"]
            if base is not None and t["cost_formula"] == "standard" and abs((base or 0) - ln["h"]) > 1e-7 and ln["h"]:
                want(tid, "base_unit_cost", ln["h"], label)
        elif ln["h_model"] == "pack":
            factor = ln["h"]["factor"]
            target_unit = ln["h"]["pack_price"] / ln["h"]["div"]
            if abs(wastage - factor) > 1e-7:
                want(tid, "wastage_factor", factor, label)
            coverage = params.get("coverage_rate", t["coverage_rate"])
            pk = params.get("product_key")
            if pk and pk in materials:
                target_cat = target_unit * coverage
                prev = catalog_targets.get(pk)
                if prev and abs(prev[0] - target_cat) > 1e-7:
                    prev[1].append(f"CONFLICT {label} wants {target_cat}")
                elif abs(materials[pk] - target_cat) > 1e-7:
                    catalog_targets.setdefault(pk, (target_cat, []))[1].append(label)
            elif params.get("cost_per_coverage_unit") is not None:
                if abs(params["cost_per_coverage_unit"] - target_unit) > 1e-7:
                    param_updates[tid] = target_unit
                    comments.setdefault(tid, label)
            elif t["base_unit_cost"] is not None and abs(t["base_unit_cost"] - target_unit) > 1e-7:
                want(tid, "base_unit_cost", target_unit, label)
        if ln["n_model"] == "per_unit" and t["labour_rate_per_unit"] is not None:
            if abs(t["labour_rate_per_unit"] - ln["n"]) > 1e-7 and t["cost_formula"] not in ("skip_hire", "tiered_disposal"):
                want(tid, "labour_rate_per_unit", round(ln["n"], 10), label)

    S = ["-- Rates-audit corrections generated by parity/audit/rates_audit.py",
         "-- Sets each value at its true source so engine-effective == workbook exactly.",
         "-- Apply ONLY as a harness-gated fix batch; re-run the full parity suite after.",
         "BEGIN;"]
    for tid, cols in sorted(tmpl_updates.items()):
        sets = ", ".join(f"{c} = {v:.10f}".rstrip("0").rstrip(".") if isinstance(v, float) else f"{c} = {v}"
                         for c, v in cols.items())
        S.append(f"-- {comments.get(tid, '')}")
        S.append(f"UPDATE costing_line_templates SET {sets} WHERE id = '{tid}';")
    for tid, v in sorted(param_updates.items()):
        S.append(f"-- {comments.get(tid, '')}: params.cost_per_coverage_unit precision")
        S.append("UPDATE costing_line_templates SET formula_params = jsonb_set(formula_params, "
                 f"'{{cost_per_coverage_unit}}', to_jsonb({v:.10f}::numeric)) WHERE id = '{tid}';")
    for pk, (target, labels) in sorted(catalog_targets.items()):
        conflicts = [l for l in labels if l.startswith("CONFLICT")]
        S.append(f"-- catalog {pk}: {'; '.join(labels[:3])}")
        if conflicts:
            S.append(f"-- SKIPPED — conflicting targets across lines: {conflicts}")
        else:
            S.append(f"UPDATE materials_catalog SET unit_cost = {target:.10f} WHERE product_key = '{pk}';")
    S.append("COMMIT;")
    (AUDIT / "proposed-corrections.sql").write_text("\n".join(S) + "\n")
    n = len(tmpl_updates) + len(param_updates) + len(catalog_targets)
    print(f"proposed corrections: {len(tmpl_updates)} template rows, {len(param_updates)} param precisions, "
          f"{len(catalog_targets)} catalog prices ({n} objects)")


def main():
    templates, materials, config = build_template_index()
    (AUDIT / "raw").mkdir(parents=True, exist_ok=True)
    all_findings, summary, unmatched_lines, matched_ids = [], [], [], set()
    all_pairs = []

    for survey_type, fname in WORKBOOKS:
        path = WB_DIR / fname
        print(f"parsing {fname} ...", flush=True)
        lines, rates = extract_workbook(path)
        (AUDIT / "raw" / f"{survey_type}.json").write_text(json.dumps(
            {"workbook": fname, "rates": rates, "lines": lines}, indent=1, default=str))
        matches, used = match_templates(lines, templates, survey_type)
        matched_ids |= used
        for ln in lines:
            ts = matches.get(ln["row"])
            if ts and len(ts) == 1:
                diff_line(ln, ts[0], materials, config, all_findings)
                all_pairs.append((survey_type, ln, ts[0]))
            elif ts:
                # one workbook row priced by MULTIPLE platform templates
                key = "+".join(t["line_key"] for t in ts)
                if ln["h_model"] in ("const", "pack"):
                    wb_unit = ln["h"]["unit_cost"] if ln["h_model"] == "pack" else ln["h"]
                    eff_sum = sum(engine_effective_material(t, materials, config)[0] for t in ts)
                    if abs(eff_sum - wb_unit) > TOL:
                        cls = ("TRUNCATED" if abs(eff_sum - wb_unit) <= TRUNC_TOL * max(1.0, abs(wb_unit))
                               else "WRONG_RATE")
                        all_findings.append({"workbook_row": ln["row"], "label": ln["label"],
                                             "template": f"{survey_type}/*/{key}", "template_id": None,
                                             "class": cls, "field": "material_unit_cost(sum)",
                                             "workbook": wb_unit, "platform": eff_sum, "note": "multi-template row"})
                else:
                    all_findings.append({"workbook_row": ln["row"], "label": ln["label"],
                                         "template": f"{survey_type}/*/{key}", "template_id": None,
                                         "class": "SPECIAL", "field": "multi_template_formula",
                                         "workbook": str(ln["h"])[:80], "platform": key,
                                         "note": "bespoke formula split across templates — verify via parity scenario"})
            elif not is_zero_feeder(ln):
                unmatched_lines.append({"workbook": survey_type, "row": ln["row"],
                                        "label": ln["label"], "h_model": ln["h_model"]})
        # global rates check
        for cfg_key, wb_key in [("hourly_labour_rate", "hourly_labour_rate"),
                                ("contractor_hourly_rate", "contractor_hourly_rate"),
                                ("vehicle_cost_per_mile", "vehicle_cost_per_mile")]:
            wv, pv = rates.get(wb_key), config.get(cfg_key)
            if wv is not None and pv is not None and abs(wv - pv) > TOL:
                all_findings.append({"workbook_row": rates.get("labour_row"),
                                     "label": f"GLOBAL {cfg_key}", "template": "pricing_config",
                                     "template_id": None, "class": "WRONG_RATE", "field": cfg_key,
                                     "workbook": wv, "platform": pv, "note": survey_type})
        summary.append({"workbook": survey_type, "lines": len(lines),
                        "matched": len(matches), "rates": rates})

    unmatched_templates = [
        {"template": f"{t['survey_type']}/{t['section_key']}/{t['line_key']}",
         "description": t["description"], "formula": t["cost_formula"]}
        for t in templates if t["id"] not in matched_ids and t.get("is_active", True)
    ]

    # ---------- report ----------
    order = ["WRONG_RATE", "MODEL_MISMATCH", "MISSING_WASTAGE", "EXTRA_WASTAGE",
             "WRONG_COVERAGE", "MARKUP", "TRUNCATED", "QUIRK", "SPECIAL"]
    by_class = {c: [f for f in all_findings if f["class"] == c] for c in order}
    L = ["# Rates Audit — workbooks vs platform templates", ""]
    L.append("Workbook formulas read directly from the .xlsm files; platform values follow the")
    L.append("engine's own precedence (catalog → params → base, × wastage_factor).")
    L.append("")
    L.append("| Workbook | line rows | matched to templates | labour rate | contractor | vehicle/mile |")
    L.append("|---|---:|---:|---:|---:|---:|")
    for s in summary:
        r = s["rates"]
        L.append(f"| {s['workbook']} | {s['lines']} | {s['matched']} | "
                 f"{r.get('hourly_labour_rate')} | {r.get('contractor_hourly_rate')} | {r.get('vehicle_cost_per_mile')} |")
    L.append("")
    L.append("| Class | Count |")
    L.append("|---|---:|")
    for c in order:
        L.append(f"| {c} | {len(by_class[c])} |")
    L.append(f"| UNMATCHED workbook lines | {len(unmatched_lines)} |")
    L.append(f"| UNMATCHED platform templates | {len(unmatched_templates)} |")
    L.append("")
    for c in order:
        if not by_class[c]:
            continue
        L.append(f"## {c} ({len(by_class[c])})")
        L.append("")
        L.append("| Workbook line | Template | Field | Workbook | Platform | Note |")
        L.append("|---|---|---|---|---|---|")
        for f in by_class[c]:
            L.append(f"| R{f['workbook_row']} {f['label'][:44]} | {f['template']} | {f['field']} | "
                     f"{f['workbook']} | {f['platform']} | {f['note']} |")
        L.append("")
    if unmatched_lines:
        L.append(f"## Unmatched workbook lines ({len(unmatched_lines)}) — no template found")
        L.append("")
        L.append("| Workbook | Row | Label | H model |")
        L.append("|---|---|---|---|")
        for u in unmatched_lines:
            L.append(f"| {u['workbook']} | R{u['row']} | {u['label'][:60]} | {u['h_model']} |")
        L.append("")
    if unmatched_templates:
        L.append(f"## Unmatched platform templates ({len(unmatched_templates)}) — no workbook line matched")
        L.append("")
        L.append("| Template | Description | Formula |")
        L.append("|---|---|---|")
        for u in unmatched_templates:
            L.append(f"| {u['template']} | {u['description'][:60]} | {u['formula']} |")
        L.append("")
    (AUDIT / "RATES_AUDIT.md").write_text("\n".join(L))

    # ---------- proposed corrections (source-precise, idempotent) ----------
    generate_corrections(all_pairs, materials)

    print(json.dumps({c: len(v) for c, v in by_class.items()}, indent=1))
    print(f"unmatched workbook lines: {len(unmatched_lines)}, unmatched templates: {len(unmatched_templates)}")
    print(f"report: {AUDIT / 'RATES_AUDIT.md'}")


if __name__ == "__main__":
    sys.exit(main())
